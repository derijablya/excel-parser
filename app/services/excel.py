from collections import defaultdict
from io import BytesIO

from fastapi import Depends, File
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app import models
from app.repository.postgres import get_session
from app.services.utils import parse_data


class ExcelService:
    def __init__(self, repository: AsyncSession = Depends(get_session)):
        self.repository = repository
        self.project = models.Project
        self.data = models.Data
        self.version = models.Version

    async def load_excel_file(self, file: File):
        data = await parse_data(file)
        version = self.version()
        self.repository.add(version)
        await self.repository.flush()
        for project, values in data.items():
            project = self.project(
                code=project[0], name=project[1], version_id=version.id
            )
            self.repository.add(project)
            await self.repository.flush()
            for value in values:
                data = self.data(
                    date=value["date"],
                    plan=value["plan_val"],
                    factual=value["fact_val"],
                    project_id=project.id,
                )
                self.repository.add(data)
        await self.repository.commit()
        return {"status": "ok"}

    async def get_excel_file(self, version):
        workbook = Workbook()
        sheet = workbook.active

        stmt = (
            select(self.version)
            .where(self.version.version == version)
            .options(selectinload(self.version.project).selectinload(self.project.data))
        )

        result = await self.repository.execute(stmt)

        version = result.scalars().first()

        unique_dates = sorted(
            {row.date for project in version.project for row in project.data}
        )

        dates_header = [date.strftime("%m/%d/%Y") for date in unique_dates]
        sheet.append(["", ""] + [date for date in dates_header for _ in range(2)])
        for col in range(1, sheet.max_column, 2):
            col_start_letter = get_column_letter(col)
            col_end_letter = get_column_letter(col + 1)
            sheet.merge_cells(f"{col_start_letter}{1}:{col_end_letter}{1}")

        headers = ["Код", "Наименование проекта"]
        headers.extend(["план", "факт"] * len(unique_dates))
        sheet.append(headers)

        data_dict = defaultdict(list)
        for project in version.project:
            for date in unique_dates:
                plan = next(
                    (row.plan for row in project.data if row.date == date),
                    None,
                )
                fact = next(
                    (row.factual for row in project.data if row.date == date),
                    None,
                )

                data_dict[(project.code, project.name)].extend([plan, fact])

        for (code, name), values in data_dict.items():
            sheet.append([code, name] + values)

        excel_data = BytesIO()
        workbook.save(excel_data)
        excel_data.seek(0)
        return excel_data
